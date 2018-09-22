import re
import dns
from dns import resolver
import socket
import smtplib

from openpyxl import load_workbook

wb = load_workbook('test1.xlsx', data_only=True)
sh = wb["Sheet1"]

for row in sh['A{}:A{}'.format(sh.min_row + 1, sh.max_row)]:
    for cell in row:
        try:
            wb.save('test1.xlsx')
            addressToVerify = cell.value
            match = re.match('^[_a-z0-9-]+(\.[_a-z0-9-]+)*@[a-z0-9-]+(\.[a-z0-9-]+)*(\.[a-z]{2,4})$', addressToVerify)
            if match == None:
                print('Bad Syntax for ' + addressToVerify)

            resolver = dns.resolver.Resolver()
            records = dns.resolver.query('lambdadirect.com', 'MX')

            mxRecord = records[0].exchange
            mxRecord = str(mxRecord)

            # Get local server hostname
            host = socket.gethostname()

            # SMTP lib setup (use debug level for full output)
            server = smtplib.SMTP()
            server.set_debuglevel(0)

            # SMTP Conversation
            server.connect(mxRecord)
            server.helo(host)
            server.mail('me@domain.com')
            print(addressToVerify)
            code, message = server.rcpt(str(addressToVerify))
            server.quit()
            print(code)
            # Assume 250 as Success
            # Assume 550 as Failure
            if code == 550:
                sh.cell(row=cell.row, column=2).value = "Fail"

            if code == 250:
                sh.cell(row=cell.row, column=2).value = "Success"

            else:
                sh.cell(row=cell.row, column=2).value = "Fail"

        except Exception as e:
            print("error",e," for address ",addressToVerify)

wb.save('test1.xlsx')
print("Done")
